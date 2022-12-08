from openpyxl import load_workbook
import pandas as pd
from itertools import islice
from flashtext import KeywordProcessor
from unidecode import unidecode

# wb = load_workbook(r'path\test.xlsx') # Load in the workbook
# class ExcelScraper(JobboardScraper):
wb = load_workbook(r'path\SelectedCompanies - Top 50.xlsx')
sheet = wb['Sheet1']
df = pd.DataFrame(sheet.values)     # Convert Sheet to DataFrame
data = sheet.values                 # Put the sheet values in `data`
cols = next(data)[1:]               # Indicate the columns in the sheet values
data = list(data)                   # Convert your data to a list
idx = [r[0] for r in data]          # Read in the data at index 0 for the indices
data = (islice(r, 1, None) for r in data)   # Slice the data at index 1
df = pd.DataFrame(data, index=idx, columns=cols)    # Make your DataFrame
# print(df)
num = 1
for nums, row in enumerate(df.itertuples(), 1): 
    # full_description = getattr(row, 'Description')
    full_description = df.at[int(nums), 'Description']
    
    skills = ''
    skills_path = r'path\skills.txt'
    with open(skills_path, 'r', encoding='utf8') as f:
        skills_list = f.read()
        skills_list = skills_list.split('\n')
        skills_keyword_processor_short = KeywordProcessor(case_sensitive=True)
        skills_keyword_processor_long = KeywordProcessor(case_sensitive=False)
        for skill in skills_list:
            if len(skill.strip()) > 4:
                skills_keyword_processor_long.add_keyword(skill.strip())
            else:
                skills_keyword_processor_short.add_keyword(skill.strip())

    if skills:
        skills = skills.replace('"', '').replace("'", '').replace('’', '').replace('`', '').strip()[:4000]
    else:
        skills = ''
        try:
            skills = ';'.join(list(set(skills_keyword_processor_short.extract_keywords(unidecode(full_description.replace('i.d.R.', ''))))) + list(set(skills_keyword_processor_long.extract_keywords(full_description))))
        except:
            None
        skills = skills[:4000]
        if len(skills) == 0:
            skills = None

    print(num)

    num += 1
    # try:
    #     df['Skills'] = df['Skills'].replace(getattr(row, 'Skills'), skills)
    df.at[int(nums), 'Skills'] = skills
df.to_excel('skills_updated3.xlsx') 
