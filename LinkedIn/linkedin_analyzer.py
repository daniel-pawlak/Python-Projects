import json
import pyodbc
import re
import traceback
from difflib import SequenceMatcher
from company_names_utilities import *
from openpyxl import Workbook


with open(r'path\db_config.json', 'r', encoding='utf-8') as f:
    db_config = json.load(f)

with open(r'path\competitors_names.json', 'r', encoding='utf-8') as f:
    competitors_names_dict = json.load(f)

with open(r'path\forbes_2000plus.txt', 'r', encoding='utf-8') as f:
    company_names = f.read().split('\n')

with pyodbc.connect('DRIVER='+db_config['driver']+';SERVER='+db_config['server']+';PORT=1433;DATABASE='+db_config['database']+';UID='+db_config['username']+';PWD='+db_config['password']) as cnxn:
    cursor = cnxn.cursor()
    cursor.execute('SELECT * FROM li_profiles_2020_2')
    all_profiles = cursor.fetchall()


rpo_list = ['ATS', 'Applicant Tracking System', 'Taleo', 'Jobvite', 'iCIMS', 'SuccessFactors', 'Success Factors', 'PeopleFluent', 'People Fluent', 'SilkRoad', 'Silk Road', 'Newton', 'Avature', 'Bullhorn', 'Sourcing', 'Screening']
rpo_list_2 = ['RPO', 'Recruitment Process Outsourcing']
msp_list = ['Contingent Workforce Solution', 'Contingent Workforce Program', 'CWS', 'CWP', 'Contractor Management', 'VMS', 'Beeline', 'IQN', 'Fieldglass', 'Kenexa', 'Services Procurement', 'Vendor Management', 'SOW']
msp_list_2 = ['MSP', 'Managed Service Provider']
triggerwords = ['via', 'through', 'for', 'on behalf', 'contracted', 'supported', '@', 'onsite', 'on site', 'on-site', 'powered by', 'secondment', 'seconded', 'remote', 'outsourced', 'client-site', 'client site']
difficult_companies = ['Total', 'Target', 'Nationwide', 'Workday', 'Bid', 'IBM', 'SAP']


profiles_to_analyze_manually = []
analyzed_profiles = []
parsed_data_dict = {}
msp_rpo_dict = {}
for profile_nr, profile in enumerate(all_profiles[29000:]):
    added_to_excel = False
    print('Profile number:', profile_nr)
    main_title = str(profile[4])
    main_location = str(profile[5])
    pos_0_title = str(profile[9])
    pos_0_company = str(profile[10])
    pos_0_date = str(profile[11])
    pos_0_desc = str(profile[14])
    pos_1_title = str(profile[15])
    pos_1_company = str(profile[16])
    pos_1_date = str(profile[17])
    pos_1_desc = str(profile[20])

    already_saved = []

    if len(str(profile[5]).split(', ')) == 3:
        main_location = str(profile[5]).split(', ')[-1]
    elif len(str(profile[5]).split(', ')) == 2:
        main_location = str(profile[5]).split(', ')[-1]
    else:
        main_location = str(profile[5])

    competitor_name = ''
    for c_name, alternative_names in competitors_names_dict.items():
        if profile[17]:
            if 'Present' in pos_1_date:
                if c_name in f'{pos_1_title} {pos_1_company}' or any([n in f'{pos_1_title} {pos_1_company}' for n in alternative_names]):
                    competitor_name = str(c_name)
        if c_name in f'{main_title} {pos_0_title} {pos_0_company}' or any([n in f'{main_title} {pos_0_title} {pos_0_company}' for n in alternative_names]):
            competitor_name = str(c_name)

    if competitor_name not in parsed_data_dict:
        parsed_data_dict[competitor_name] = {}
    else:
        if main_location not in parsed_data_dict[competitor_name]:
            parsed_data_dict[competitor_name][main_location] = {}

    if competitor_name not in msp_rpo_dict:
        msp_rpo_dict[competitor_name] = {}
    else:
        if main_location not in msp_rpo_dict[competitor_name]:
            msp_rpo_dict[competitor_name][main_location] = {'msp': 0, 'rpo': 0, 'msp_profiles': 0, 'rpo_profiles': 0}

    # if company not in already_saved:    
    #     if company not in parsed_data_dict[competitor_name][main_location]:
    #         parsed_data_dict[competitor_name][main_location][company] = 1
    #     else:
    #         parsed_data_dict[competitor_name][main_location][company] += 1

    try:
        # MSP RPO
        rpo_points = 0
        msp_points = 0
        if profile[17]:
            if 'Present' in pos_1_date:
                for msp in msp_list:
                    msp_points += len(re.findall(r'\b' + msp + r'\b', f"{pos_1_title} {pos_1_company} {pos_1_desc}", re.IGNORECASE))
                for msp in msp_list_2:
                    msp_points += len(re.findall(r'\b' + msp + r'\b', f"{pos_1_title} {pos_1_company} {pos_1_desc}", re.IGNORECASE)) * 2
                for rpo in rpo_list:
                    rpo_points += len(re.findall(r'\b' + rpo + r'\b', f"{pos_1_title} {pos_1_company} {pos_1_desc}", re.IGNORECASE))
                for rpo in rpo_list_2:
                    rpo_points += len(re.findall(r'\b' + rpo + r'\b', f"{pos_1_title} {pos_1_company} {pos_1_desc}", re.IGNORECASE)) * 2
        for msp in msp_list:
            msp_points += len(re.findall(r'\b' + msp + r'\b', f"{main_title} {pos_0_title} {pos_0_company} {pos_0_desc}", re.IGNORECASE))
        for msp in msp_list_2:
            msp_points += len(re.findall(r'\b' + msp + r'\b', f"{main_title} {pos_0_title} {pos_0_company} {pos_0_desc}", re.IGNORECASE)) * 2
        for rpo in rpo_list:
            rpo_points += len(re.findall(r'\b' + rpo + r'\b', f"{main_title} {pos_0_title} {pos_0_company} {pos_0_desc}", re.IGNORECASE))
        for rpo in rpo_list_2:
            rpo_points += len(re.findall(r'\b' + rpo + r'\b', f"{main_title} {pos_0_title} {pos_0_company} {pos_0_desc}", re.IGNORECASE)) * 2
        msp_rpo_dict[competitor_name][main_location]['msp'] += msp_points
        msp_rpo_dict[competitor_name][main_location]['rpo'] += rpo_points
        if msp_points > 0:
            msp_rpo_dict[competitor_name][main_location]['msp_profiles'] += 1
        if rpo_points > 0:
            msp_rpo_dict[competitor_name][main_location]['rpo_profiles'] += 1


        # Triggerword and other company name in main title or present position title/company name
        if profile[17]:
            if 'Present' in pos_1_date:
                for profile_part in [main_title, pos_0_title, pos_0_company, pos_1_title, pos_1_company]:
                    for company in company_names:
                        if len(company) > 3:
                            if any(re.search(r'\b' + word + r'\b', profile_part, re.IGNORECASE) for word in triggerwords) and re.search(r'\b' + company + r'\b', profile_part, re.IGNORECASE):
                                print(company, '-->', profile_part)
                                print('1----------------------')
                                if company not in already_saved:
                                    already_saved.append(company)
                                    if company not in parsed_data_dict[competitor_name][main_location]:
                                        parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                        added_to_excel = True
                                    else:
                                        parsed_data_dict[competitor_name][main_location][company][0] += 1
                                        parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                        added_to_excel = True
                        else:
                            if any(re.search(r'\b' + word + r'\b', profile_part, re.IGNORECASE) for word in triggerwords) and re.search(r'\b' + company + r'\b', profile_part):
                                print(company, '-->', profile_part)
                                print('1----------------------')
                                if company not in already_saved:
                                    already_saved.append(company)
                                    if company not in parsed_data_dict[competitor_name][main_location]:
                                        parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                        added_to_excel = True
                                    else:
                                        parsed_data_dict[competitor_name][main_location][company][0] += 1
                                        parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                        added_to_excel = True
        for company in company_names:
            if len(company) > 3:
                for profile_part in [main_title, pos_0_title, pos_0_company]:
                    if any(re.search(r'\b' + word + r'\b', profile_part, re.IGNORECASE) for word in triggerwords) and re.search(r'\b' + company + r'\b', profile_part, re.IGNORECASE):
                        print(company, '-->', profile_part)
                        print('2----------------------')
                        if company not in already_saved:
                            already_saved.append(company)
                            if company not in parsed_data_dict[competitor_name][main_location]:
                                parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                added_to_excel = True
                            else:
                                parsed_data_dict[competitor_name][main_location][company][0] += 1
                                parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                added_to_excel = True
            else:
                for profile_part in [main_title, pos_0_title, pos_0_company]:
                    if any(re.search(r'\b' + word + r'\b', profile_part, re.IGNORECASE) for word in triggerwords) and re.search(r'\b' + company + r'\b', profile_part):
                        print(company, '-->', profile_part)
                        print('2----------------------')
                        if company not in already_saved:
                            already_saved.append(company)
                            if company not in parsed_data_dict[competitor_name][main_location]:
                                parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                added_to_excel = True
                            else:
                                parsed_data_dict[competitor_name][main_location][company][0] += 1
                                parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                added_to_excel = True
        # Different companies for the same period of time
        if profile[17]:
            if 'Present' in pos_0_date and 'Present' in pos_1_date and any(c in pos_1_company for c in competitors_names_dict[competitor_name] + [competitor_name]) and any(c not in pos_0_company for c in competitors_names_dict[competitor_name] + [competitor_name]):
                company = pos_0_company
                print(company)
                print('3----------------------')
                if company and company not in already_saved:
                    already_saved.append(company)
                    if company not in parsed_data_dict[competitor_name][main_location]:
                        parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                        added_to_excel = True
                    else:
                        parsed_data_dict[competitor_name][main_location][company][0] += 1
                        parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                        added_to_excel = True
        # Triggerwords and other company name in present position description
        if profile[17]:
            if 'Present' in pos_1_date:
                for triggerword in triggerwords:
                    if re.search(r'\b' + triggerword + r'\b', pos_0_desc + ' ' + pos_1_desc):
                        start_index = re.search(r'\b' + triggerword + r'\b', pos_0_desc + ' ' + pos_1_desc).start() - 50
                        end_index = re.search(r'\b' + triggerword + r'\b', pos_0_desc + ' ' + pos_1_desc).end() + 50
                        description_chunk = (pos_0_desc + ' ' + pos_1_desc)[start_index : end_index]
                        for company in company_names:
                            if len(company) > 3:
                                if re.search(r'\b' + company + r'\b', description_chunk, re.IGNORECASE):
                                    print(company, '-->', description_chunk)
                                    print('4----------------------')
                                    if company not in already_saved:
                                        already_saved.append(company)
                                        if company not in parsed_data_dict[competitor_name][main_location]:
                                            parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                            added_to_excel = True
                                        else:
                                            parsed_data_dict[competitor_name][main_location][company][0] += 1
                                            parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                            added_to_excel = True
                            else:
                                if re.search(r'\b' + company + r'\b', description_chunk):
                                    print(company, '-->', description_chunk)
                                    print('4----------------------')
                                    if company not in already_saved:
                                        already_saved.append(company)
                                        if company not in parsed_data_dict[competitor_name][main_location]:
                                            parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                            added_to_excel = True
                                        else:
                                            parsed_data_dict[competitor_name][main_location][company][0] += 1
                                            parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                            added_to_excel = True
        for triggerword in triggerwords:
            if re.search(r'\b' + triggerword + r'\b', str(profile[14])):
                start_index = re.search(r'\b' + triggerword + r'\b', pos_0_desc).start() - 50
                end_index = re.search(r'\b' + triggerword + r'\b', pos_0_desc).end() + 50
                description_chunk = pos_0_desc[start_index : end_index]
                for company in company_names:
                    if len(company) > 3:
                        if re.search(r'\b' + company + r'\b', description_chunk, re.IGNORECASE):
                            print(company, '-->', description_chunk)
                            print('5----------------------')
                            if company not in already_saved:
                                already_saved.append(company)
                                if company not in parsed_data_dict[competitor_name][main_location]:
                                    parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                    added_to_excel = True
                                else:
                                    parsed_data_dict[competitor_name][main_location][company][0] += 1
                                    parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                    added_to_excel = True
                    else:
                        if re.search(r'\b' + company + r'\b', description_chunk):
                            print(company, '-->', description_chunk)
                            print('5----------------------')
                            if company not in already_saved:
                                already_saved.append(company)
                                if company not in parsed_data_dict[competitor_name][main_location]:
                                    parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                    added_to_excel = True
                                else:
                                    parsed_data_dict[competitor_name][main_location][company][0] += 1
                                    parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                    added_to_excel = True
        # Company name core in main title or present position title/description
        if profile[17]:
            if 'Present' in profile[17]:
                for profile_part in [main_title, pos_0_company, pos_0_title, pos_0_desc, pos_1_company, pos_1_title, pos_0_desc]:
                    for company in company_names:
                        if re.search(r'\b' + get_core_name(company) + r'\b', profile_part, re.IGNORECASE):
                            print(company, '-->', profile_part)
                            print('7----------------------')
                            if company not in already_saved:
                                already_saved.append(company)
                                if company not in parsed_data_dict[competitor_name][main_location]:
                                    parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                    added_to_excel = True
                                else:
                                    parsed_data_dict[competitor_name][main_location][company][0] += 1
                                    parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                    added_to_excel = True
            else:
                for profile_part in [main_title, pos_0_company, pos_0_title, pos_0_desc]:
                    for company in company_names:
                        if re.search(r'\b' + get_core_name(company) + r'\b', profile_part, re.IGNORECASE):
                            print(company, '-->', profile_part)
                            print('7----------------------')
                            if company not in already_saved:
                                already_saved.append(company)
                                if company not in parsed_data_dict[competitor_name][main_location]:
                                    parsed_data_dict[competitor_name][main_location][company] = [1, str(profile[3])]
                                    added_to_excel = True
                                else:
                                    parsed_data_dict[competitor_name][main_location][company][0] += 1
                                    parsed_data_dict[competitor_name][main_location][company][1] = parsed_data_dict[competitor_name][main_location][company][1] + f' {str(profile[3])}'
                                    added_to_excel = True
        if not added_to_excel:
            # Triggerword in main title or present position title/company name
            if profile[17]:
                if 'Present' in profile[17]:
                    for profile_part in [main_title, pos_0_title, pos_0_company, pos_1_title, pos_1_company]:
                        if any(re.search(r'\b' + word + r'\b', profile_part, re.IGNORECASE) for word in triggerwords):
                            print('triggerword', '-->', profile_part)
                            print('6----------------------')
                            profiles_to_analyze_manually.append((str(profile[3]), main_location, 'triggerword', profile_part))
                else:
                    for profile_part in [main_title, pos_0_title, pos_0_company]:
                        if any(re.search(r'\b' + word + r'\b', profile_part, re.IGNORECASE) for word in triggerwords):
                            print('triggerword', '-->', profile_part)
                            print('6----------------------')
                            profiles_to_analyze_manually.append((str(profile[3]), main_location, 'triggerword', profile_part))
    except Exception as e:
        print(e)
    print('=====================================')

wb = Workbook()
ws = wb.active
num = 1
for competitor, locations in parsed_data_dict.items():
    for location, clients in locations.items():
        for client, profiles_num in clients.items():
            ws['A' + str(num)] = competitor
            ws['B' + str(num)] = location
            ws['C' + str(num)] = client
            ws['D' + str(num)] = profiles_num[0]
            ws['E' + str(num)] = profiles_num[1]
            num += 1
wb.save("linkedin_profiles_2020_q4_countries_6.xlsx")

wb = Workbook()
ws = wb.active
num = 1
for profile in profiles_to_analyze_manually:
    ws['A' + str(num)] = profile[0]
    ws['B' + str(num)] = profile[1]
    ws['C' + str(num)] = profile[2]
    ws['D' + str(num)] = profile[3]
    num += 1
wb.save("linkedin_profiles_2020_q4_to_analyze_6.xlsx")

wb = Workbook()
ws = wb.active
num = 1
for competitor, locations in msp_rpo_dict.items():
    for location in locations:
        ws['A' + str(num)] = competitor
        ws['B' + str(num)] = location
        ws['C' + str(num)] = msp_rpo_dict[competitor][location]['msp']
        ws['D' + str(num)] = msp_rpo_dict[competitor][location]['rpo']
        ws['E' + str(num)] = msp_rpo_dict[competitor][location]['msp_profiles']
        ws['F' + str(num)] = msp_rpo_dict[competitor][location]['rpo_profiles']
        num += 1
wb.save("linkedin_profiles_2020_q4_msp_rpo_6.xlsx")
